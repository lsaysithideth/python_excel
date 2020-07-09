import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows #import module to convert dataframe that is usable by openpyxl

wb = load_workbook('regions.xlsx')
ws = wb.active
df = pd.read_excel('all_shifts.xlsx')
df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
df1['Total'] = df1['Cost per'] * df1['Units Sold']

#use for loop to add to excel file worksheet
rows = dataframe_to_rows(df1, index=False)
for r_idx, row in enumerate(rows,1): #iterates like a regular for loop but also keeping record of the row and column indices
    for c_idx, col in enumerate(row,6): #to not overwrite current data start at column 6
        ws.cell(row=r_idx, column=c_idx, value=col)

wb.save('combined.xlsx')