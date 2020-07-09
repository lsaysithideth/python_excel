import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')
df_all = pd.concat([df_1,df_2,df_3], sort=False) #make sure sort=False so columns stay in the same order

to_excel = df_all.to_excel('all_shifts_new.xlsx', index = None)

wb = load_workbook('all_shifts_new.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)

ws = wb.active
df = pd.read_excel('all_shifts.xlsx')
df['Total'] = df['Cost per'] * df['Units Sold']
print(df)

wb.save('totaled_new.xlsx')