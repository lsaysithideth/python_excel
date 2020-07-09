import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')
df_all = pd.concat([df_1,df_2,df_3], sort=False) #make sure sort=False so columns stay in the same order

to_excel = df_all.to_excel('all_shifts_new.xlsx', index = None)

wb = load_workbook('all_shifts_new.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

e_col, f_col = ['E', 'F']
for row in range(2,300): #don't include first row which contains headers so start at 2
    result_cell = 'G{}'.format(row) #grab results cell, all cells in G column
    #next grab other values from other columns stored
    e_value = ws[e_col + str(row)].value #index worksheet by e column string value and string cast to row value
    f_value = ws[f_col + str(row)].value
    ws[result_cell] = e_value *f_value #each time loop iterates move one row down in current column

wb.save('totaled.xlsx')