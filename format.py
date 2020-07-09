from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side, NamedStyle
#from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

for i in range(1,20):
        ws.append(range(300))

ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

cell = ws['B2']
cell.font = Font(color=colors.COLOR_INDEX[2], size=20, italic=True)

"""
# Will remove these definitions in a future release
BLACK = COLOR_INDEX[0]
WHITE = COLOR_INDEX[1]
#RED = COLOR_INDEX[2]
#DARKRED = COLOR_INDEX[8]
BLUE = COLOR_INDEX[4]
#DARKBLUE = COLOR_INDEX[12]
#GREEN = COLOR_INDEX[3]
#DARKGREEN = COLOR_INDEX[9]
#YELLOW = COLOR_INDEX[5]
#DARKYELLOW = COLOR_INDEX[19]

https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/colors.html

"""

cell.value = 'Merged Cell'
cell.alignment = Alignment(horizontal='right', vertical='bottom')
cell.fill = GradientFill(stop=("000000","FFFFFF"))
wb.save('text.xlsx')

highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick',color='000000')
highlight.border = Border(left=bd,top=bd,right=bd,bottom=bd)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

count = 0 
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count = count + 1 #moves one row down automatically in loop
wb.save('highlight.xlsx')
